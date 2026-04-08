"""
Excel 読み込み・フィルタリングモジュール
=========================================
excel_md_tool (useExcel.ts) と同等のロジックを Python / openpyxl で実装。

対応仕様:
  - ヘッダー行の範囲指定（複数行ヘッダーは " / " 結合、セル内改行は除去）
  - データ開始行の指定
  - 列範囲の指定（A, B, C ... の列ラベル形式）
  - フィルター条件（列名＋値の AND フィルター、value: 単一値 / values: 複数値 OR）
  - 日付セルは "YYYY/MM/DD" 文字列に変換
  - 空行はスキップ
"""

from __future__ import annotations

import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.utils import column_index_from_string
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        _RICH_TEXT_AVAILABLE = True
    except ImportError:
        _RICH_TEXT_AVAILABLE = False
except ImportError:
    raise ImportError(
        "openpyxl が必要です。`pip install openpyxl` を実行してください。"
    )


# ------------------------------------------------------------------
# ユーティリティ
# ------------------------------------------------------------------

def col_letter_to_index(letter: str) -> int:
    """列ラベル（"A", "B", ..., "AA"）を 0 始まりの整数インデックスに変換"""
    return column_index_from_string(letter.upper()) - 1


def cell_to_str(value: Any) -> str:
    """セル値を文字列に変換（None/空は空文字列、日付は YYYY/MM/DD）"""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        if isinstance(value, datetime):
            return value.strftime("%Y/%m/%d")
        return value.strftime("%Y/%m/%d")
    # リッチテキストはプレーンテキストとして結合して返す
    if _RICH_TEXT_AVAILABLE and isinstance(value, CellRichText):
        return "".join(
            str(run.text) if isinstance(run, TextBlock) else str(run)
            for run in value
        ).strip()
    return str(value).strip()


def cell_to_markdown(cell: Any) -> str:
    """
    セルの値と書式を Markdown 文字列に変換する。

    対応書式:
        取り消し線（strike）→ ~~text~~

    優先順位:
        1. CellRichText（リッチテキスト）→ ランごとの書式を使用（最優先）
           セル全体スタイル（cell.font.strike）より細かいランレベルの情報が
           正確なため、CellRichText の場合は run.font.strike を優先する。
           これにより「セル内の一部のみ取り消し線」が正しく反映される。

        2. cell.font.strike == True（CellRichText でないセル）→ ~~セル全体~~
           日付・プレーンテキストなど CellRichText 以外のセルに適用。

        3. 書式なし → cell_to_str() と同じ出力（後方互換）

    備考:
        CellRichText セルに cell.font.strike=True が付いていても無視する。
        Excel はセルの一部に取り消し線を設定した際にセルレベルのスタイルにも
        strike=True を記録することがあるが、正確な情報はランレベルにある。
    """
    value = cell.value

    if value is None:
        return ""

    # ---- ① リッチテキスト（セル内部分書式・最優先）----
    # CellRichText の場合はランレベルの書式を使用する。
    # cell.font.strike は「セル全体スタイル」として Excel が自動設定することがあり、
    # 実際のランレベルの取り消し線範囲と一致しないケースがあるため無視する。
    if _RICH_TEXT_AVAILABLE and isinstance(value, CellRichText):
        result = ""
        for run in value:
            if isinstance(run, TextBlock):
                text = str(run.text)
                is_struck = bool(run.font and run.font.strike)
            else:
                text = str(run)
                is_struck = False

            if not text:
                continue

            if is_struck:
                # ~~ の前: 直前の文字がスペース・改行でなければスペースを挿入
                if result and result[-1] not in (' ', '\n', '\r'):
                    result += ' '
                result += f'~~{text}~~'
            else:
                # ~~ の直後: 現テキストがスペース・改行始まりでなければスペースを挿入
                if result.endswith('~~') and text[0] not in (' ', '\n', '\r'):
                    result += ' '
                result += text

        return result.strip()

    # ---- ② セル全体に取り消し線（日付・プレーンテキスト）----
    # CellRichText 以外（日付セル・通常テキスト）のセルで
    # cell.font.strike が True の場合、値全体を ~~ で囲む。
    if cell.font and cell.font.strike:
        text = cell_to_str(value)
        return f'~~{text}~~' if text else ''

    # ---- ③ 書式なし ----
    return cell_to_str(value)


# ------------------------------------------------------------------
# ExcelReader
# ------------------------------------------------------------------

class ExcelReader:
    """
    config (sources[i].excel) に従って Excel を読み込み、
    ヘッダー名をキーにした行データのリストを返す。

    Parameters
    ----------
    excel_config : dict
        path            : str   Excel ファイルパス
        sheet           : str   シート名（省略時: 最初のシート）
        header_start_row: int   ヘッダー開始行（1始まり、デフォルト 1）
        header_end_row  : int   ヘッダー終了行（デフォルト = header_start_row）
        data_start_row  : int   データ開始行（デフォルト = header_end_row + 1）
        col_start       : str   読み込み開始列ラベル（デフォルト "A"）
        col_end         : str   読み込み終了列ラベル（デフォルト: シート最終列）
    """

    MULTI_HEADER_SEP = " / "  # 複数行ヘッダーの結合区切り文字

    def __init__(self, excel_config: dict):
        self.path = Path(excel_config["path"]).expanduser()
        self.sheet_name: str | None = excel_config.get("sheet")
        self.header_start_row: int = int(excel_config.get("header_start_row", 1))
        self.header_end_row: int = int(
            excel_config.get("header_end_row", self.header_start_row)
        )
        self.data_start_row: int = int(
            excel_config.get("data_start_row", self.header_end_row + 1)
        )
        self.col_start_str: str = str(excel_config.get("col_start", "A")).upper()
        self.col_end_str: str | None = (
            str(excel_config.get("col_end")).upper()
            if excel_config.get("col_end")
            else None
        )

        # バリデーション
        if self.header_start_row < 1:
            raise ValueError("header_start_row は 1 以上を指定してください。")
        if self.header_end_row < self.header_start_row:
            raise ValueError("header_end_row は header_start_row 以上を指定してください。")
        if self.data_start_row <= self.header_end_row:
            raise ValueError("data_start_row は header_end_row より大きい値を指定してください。")

    # ------------------------------------------------------------------

    def _load_sheet(self, rich_text: bool = False):
        """
        ワークブックを開いてシートを返す。

        Parameters
        ----------
        rich_text : bool
            True のとき rich_text=True でワークブックを開く。
            openpyxl がリッチテキストセルを CellRichText として返すようになり、
            cell_to_markdown() による書式付き変換が可能になる。
            False（デフォルト）は通常の data_only=True で開く。
        """
        if not self.path.exists():
            raise FileNotFoundError(f"Excel ファイルが見つかりません: {self.path}")

        if rich_text and _RICH_TEXT_AVAILABLE:
            wb = openpyxl.load_workbook(str(self.path), data_only=True, rich_text=True)
        else:
            wb = openpyxl.load_workbook(str(self.path), data_only=True)

        if self.sheet_name:
            if self.sheet_name not in wb.sheetnames:
                available = ", ".join(wb.sheetnames)
                raise ValueError(
                    f"シート「{self.sheet_name}」が見つかりません。"
                    f"（利用可能: {available}）"
                )
            ws = wb[self.sheet_name]
        else:
            ws = wb.active

        return ws

    def _resolve_col_range(self, ws) -> tuple[int, int]:
        """
        col_start / col_end を 0 始まりの列インデックスに解決する。
        col_end が未指定の場合はシートの最終列を使う。
        """
        col_start_idx = col_letter_to_index(self.col_start_str)

        if self.col_end_str:
            col_end_idx = col_letter_to_index(self.col_end_str)
        else:
            # シートの最終列（max_column は 1 始まり）
            col_end_idx = (ws.max_column or 1) - 1

        if col_start_idx > col_end_idx:
            raise ValueError(
                f"col_start ({self.col_start_str}) は col_end ({self.col_end_str}) より前の列を指定してください。"
            )
        return col_start_idx, col_end_idx

    def _build_headers(self, ws, col_start_idx: int, col_end_idx: int) -> list[str]:
        """
        ヘッダー行（header_start_row ～ header_end_row）を読み取り、
        複数行の場合は MULTI_HEADER_SEP で結合したヘッダー名リストを返す。
        """
        headers = []
        for col_idx in range(col_start_idx, col_end_idx + 1):
            parts = []
            for row_idx in range(self.header_start_row, self.header_end_row + 1):
                # openpyxl は 1 始まり
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                val = cell_to_str(cell.value).replace("\r", "").replace("\n", "")
                if val:
                    parts.append(val)
            headers.append(self.MULTI_HEADER_SEP.join(parts) if parts else f"Col{col_idx + 1}")
        return headers

    def _build_rows(
        self,
        ws,
        headers: list[str],
        col_start_idx: int,
        col_end_idx: int,
        use_markdown: bool = False,
    ) -> list[dict[str, str]]:
        """
        データ行（data_start_row ～ シート最終行）を読み取り、
        {ヘッダー名: セル値} の dict リストを返す。
        空行（全セルが空）はスキップする。

        Parameters
        ----------
        use_markdown : bool
            True のとき cell_to_markdown() で書式付き Markdown 文字列を生成する。
            False（デフォルト）は cell_to_str() でプレーンテキストを返す。
            空行判定はプレーンテキストで行うため、use_markdown=True のときも
            空行は正しくスキップされる。
        """
        rows = []
        max_row = ws.max_row or 0

        for row_idx in range(self.data_start_row, max_row + 1):
            row_data = {}
            is_empty = True
            for i, col_idx in enumerate(range(col_start_idx, col_end_idx + 1)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                plain = cell_to_str(cell.value)
                if plain:
                    is_empty = False
                if use_markdown:
                    row_data[headers[i]] = cell_to_markdown(cell)
                else:
                    row_data[headers[i]] = plain

            if not is_empty:
                rows.append(row_data)

        return rows

    def _build_rows_dual(
        self,
        ws,
        headers: list[str],
        col_start_idx: int,
        col_end_idx: int,
    ) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
        """
        ワークブックを1回だけ走査し、プレーンテキスト行と書式付き Markdown 行を
        同時に構築して返す。

        plain_rows と formatted_rows は常に同じ長さ・同じ順序になることが保証される。
        （_build_rows を2回呼ぶ方式では、rich_text オプションの有無による
         空行判定の違いでインデックスがズレる可能性があった）

        Returns
        -------
        plain_rows     : list[dict[str, str]]  プレーンテキスト行
        formatted_rows : list[dict[str, str]]  書式付き Markdown 行（取り消し線 等）
        """
        plain_rows: list[dict[str, str]] = []
        formatted_rows: list[dict[str, str]] = []
        max_row = ws.max_row or 0

        for row_idx in range(self.data_start_row, max_row + 1):
            plain_data: dict[str, str] = {}
            fmt_data: dict[str, str] = {}
            is_empty = True

            for i, col_idx in enumerate(range(col_start_idx, col_end_idx + 1)):
                cell = ws.cell(row=row_idx, column=col_idx + 1)
                plain = cell_to_str(cell.value)
                if plain:
                    is_empty = False
                plain_data[headers[i]] = plain
                fmt_data[headers[i]] = cell_to_markdown(cell)

            if not is_empty:
                plain_rows.append(plain_data)
                formatted_rows.append(fmt_data)

        return plain_rows, formatted_rows

    # ------------------------------------------------------------------

    def read(self) -> tuple[list[str], list[dict[str, str]]]:
        """
        Excel ファイルを読み込んでヘッダーと行データを返す。

        Returns
        -------
        headers : list[str]
            ヘッダー名リスト（複数行ヘッダーは " / " 結合）
        rows    : list[dict[str, str]]
            データ行リスト。各要素は {ヘッダー名: 値} の dict（プレーンテキスト）
        """
        ws = self._load_sheet()
        col_start_idx, col_end_idx = self._resolve_col_range(ws)
        headers = self._build_headers(ws, col_start_idx, col_end_idx)
        rows = self._build_rows(ws, headers, col_start_idx, col_end_idx)
        return headers, rows

    def read_with_format(self) -> tuple[list[str], list[dict[str, str]], list[dict[str, str]]]:
        """
        Excel ファイルを読み込んで、プレーンテキスト行と書式付き Markdown 行の両方を返す。

        書式付き行はセルの取り消し線を ~~ に変換した Markdown 文字列を持つ。
        プレーンテキスト行はフィルタリング・件名・担当者解決などに使用し、
        書式付き行は description の本文生成にのみ使用する。

        ワークブックは 1 回だけ（rich_text=True で）開き、プレーンと書式付きを
        同一パスで構築する。これにより plain_rows と formatted_rows が
        常に同じ長さ・同じ順序で対応することを保証する。

        openpyxl の rich_text オプションが利用できない場合（古いバージョン等）は
        プレーンテキストを formatted_rows としても返す（警告を出力）。

        Returns
        -------
        headers       : list[str]           ヘッダー名リスト
        plain_rows    : list[dict[str, str]] プレーンテキスト行
        formatted_rows: list[dict[str, str]] 書式付き Markdown 行
        """
        import sys

        if not _RICH_TEXT_AVAILABLE:
            print(
                "  ⚠ openpyxl のリッチテキスト機能が利用できません（rich_text オプションは無視されます）。"
                " openpyxl を最新版にアップグレードしてください。",
                file=sys.stderr,
            )
            headers, plain_rows = self.read()
            return headers, plain_rows, plain_rows

        # ワークブックを rich_text=True で 1 回だけ開く。
        # plain_rows と formatted_rows を同一パスで構築するため、
        # 空行判定が両者で必ず一致し、インデックスがズレない。
        ws = self._load_sheet(rich_text=True)
        col_start_idx, col_end_idx = self._resolve_col_range(ws)
        headers = self._build_headers(ws, col_start_idx, col_end_idx)
        plain_rows, formatted_rows = self._build_rows_dual(ws, headers, col_start_idx, col_end_idx)

        return headers, plain_rows, formatted_rows

    @staticmethod
    def filter_rows(
        rows: list[dict[str, str]],
        filters: list[dict] | None,
    ) -> list[dict[str, str]]:
        """
        AND 条件でフィルタリングして合致する行だけを返す。

        filters の各要素:
            col_name : str        対象のヘッダー名
            value    : str        一致すべき単一値（values と排他）
            values   : list[str]  一致すべき値のリスト（いずれかに一致すれば OK、OR 条件）
            match    : str        "exact"（デフォルト）/ "contains" / "startswith"
                                  value / values いずれにも適用される

        複数の filter 条件は AND で評価される。
        filters が None または空の場合は全行を返す。
        存在しない col_name はその条件をスキップする（警告は呼び出し元で出す）。
        """
        if not filters:
            return rows

        result = []
        for row in rows:
            match_all = True
            for cond in filters:
                col = cond.get("col_name", "")
                match_type = cond.get("match", "exact")

                if col not in row:
                    # 条件列がヘッダーに存在しない場合はスキップ
                    continue

                actual = row[col]

                # values（リスト）が指定されていれば OR マッチ、なければ単一 value でマッチ
                raw_values = cond.get("values")
                if raw_values is not None:
                    candidates = [str(v) for v in raw_values]
                else:
                    candidates = [str(cond.get("value", ""))]

                if match_type == "contains":
                    matched = any(c in actual for c in candidates)
                elif match_type == "startswith":
                    matched = any(actual.startswith(c) for c in candidates)
                else:  # exact
                    matched = actual in candidates

                if not matched:
                    match_all = False
                    break

            if match_all:
                result.append(row)

        return result
