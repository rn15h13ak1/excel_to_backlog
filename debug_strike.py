"""
取り消し線デバッグスクリプト
==============================
config.yaml の設定（ヘッダー行・データ開始行）を読み込んで動作します。

使い方:
    python debug_strike.py <データ行番号> <列名> [ソース名]

    データ行番号: データの何行目か（1始まり）。config の data_start_row を基準とする。
                  例: data_start_row=2 のとき、データ行番号=1 → Excel の2行目

    ソース名: config.yaml に複数ソースがある場合に指定（省略時は最初のソース）

例:
    python debug_strike.py 1 "詳細"
    python debug_strike.py 3 "対応内容" "タスク管理表"
"""

import sys
from pathlib import Path


def load_config(config_path: Path) -> dict:
    try:
        import yaml
    except ImportError:
        print("PyYAML が必要です: pip install pyyaml")
        sys.exit(1)
    with open(config_path, encoding="utf-8") as f:
        return yaml.safe_load(f)


def find_source(sources: list, source_name: str | None) -> dict:
    if not sources:
        print("config.yaml に sources が定義されていません")
        sys.exit(1)
    if source_name:
        for s in sources:
            if s.get("name") == source_name:
                return s
        print(f"ソース '{source_name}' が見つかりません")
        print(f"利用可能なソース: {[s.get('name') for s in sources]}")
        sys.exit(1)
    return sources[0]


def main():
    args = sys.argv[1:]

    # 引数またはインタラクティブ入力
    if len(args) >= 1:
        data_row_num = int(args[0])
    else:
        data_row_num = int(input("データ行番号（1始まり）: ").strip())

    if len(args) >= 2:
        col_name = args[1]
    else:
        col_name = input("確認したい列名（例: 詳細）: ").strip()

    source_name = args[2] if len(args) >= 3 else None

    # config.yaml 読み込み
    config_path = Path("config.yaml")
    if not config_path.exists():
        print(f"config.yaml が見つかりません（{Path.cwd()}）")
        sys.exit(1)

    cfg = load_config(config_path)
    source = find_source(cfg.get("sources", []), source_name)
    excel_cfg = source.get("excel", {})

    excel_path     = excel_cfg.get("path", "")
    sheet_name     = excel_cfg.get("sheet", "")
    header_start   = int(excel_cfg.get("header_start_row", 1))
    header_end     = int(excel_cfg.get("header_end_row", header_start))
    data_start     = int(excel_cfg.get("data_start_row", header_end + 1))
    rich_text_cfg  = source.get("issue_mapping", {}).get("rich_text", False)

    # データ行番号 → Excel 実際の行番号
    excel_row = data_start + (data_row_num - 1)

    print()
    print("=" * 60)
    print("【設定確認】config.yaml から読み込んだ値")
    print("=" * 60)
    print(f"  ソース名        : {source.get('name')}")
    print(f"  Excel パス      : {excel_path}")
    print(f"  シート          : {sheet_name}")
    print(f"  header_start_row: {header_start}")
    print(f"  header_end_row  : {header_end}")
    print(f"  data_start_row  : {data_start}")
    print(f"  rich_text       : {rich_text_cfg}")
    print(f"  → データ行番号 {data_row_num} = Excel {excel_row} 行目")

    print()
    print("=" * 60)
    print("【チェック1】openpyxl バージョン")
    print("=" * 60)
    import openpyxl
    print(f"openpyxl: {openpyxl.__version__}")

    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        print("CellRichText: 利用可能 ✓")
    except ImportError:
        print("CellRichText: 利用不可 ✗ → openpyxl >= 3.1.0 が必要です")
        return

    print()
    print("=" * 60)
    print("【チェック2】ファイルを rich_text=True で読み込み")
    print("=" * 60)

    try:
        wb = openpyxl.load_workbook(excel_path, rich_text=True, data_only=True)
    except TypeError:
        print("rich_text 引数非対応 ✗ → openpyxl のバージョンを上げてください")
        return
    except FileNotFoundError:
        print(f"ファイルが見つかりません: {excel_path}")
        return

    print("ファイル読み込み成功 ✓")

    if sheet_name and sheet_name not in wb.sheetnames:
        print(f"シート '{sheet_name}' が見つかりません。利用可能: {wb.sheetnames}")
        sys.exit(1)
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"シート '{ws.title}' を使用")

    print()
    print("=" * 60)
    print(f"【チェック3】ヘッダー行（{header_start}〜{header_end} 行目）から列を特定")
    print("=" * 60)

    # 複数行ヘッダーを " / " で結合（excel_reader.py と同じロジック）
    header_map: dict[str, int] = {}  # 結合ヘッダー名 → 列インデックス(1始まり)
    col_start_letter = excel_cfg.get("col_start", "A")
    col_end_letter   = excel_cfg.get("col_end", None)

    from openpyxl.utils import column_index_from_string
    col_start_idx = column_index_from_string(col_start_letter.upper())
    col_end_idx   = column_index_from_string(col_end_letter.upper()) if col_end_letter else ws.max_column

    for col_idx in range(col_start_idx, col_end_idx + 1):
        parts = []
        for r in range(header_start, header_end + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        if parts:
            combined = " / ".join(parts)
            header_map[combined] = col_idx
            print(f"  列{col_idx}: {combined}")

    print()
    print("=" * 60)
    print(f"【チェック4】列 '{col_name}' のセル詳細（Excel {excel_row} 行目）")
    print("=" * 60)

    if col_name not in header_map:
        print(f"列 '{col_name}' が見つかりません")
        print(f"利用可能な列名: {list(header_map.keys())}")
        sys.exit(1)

    col_idx = header_map[col_name]
    target_cell = ws.cell(row=excel_row, column=col_idx)
    val = target_cell.value

    print(f"Excel 位置: {target_cell.coordinate}")
    print(f"値の型    : {type(val).__name__}")
    print(f"値        : {repr(val)}")

    if isinstance(val, CellRichText):
        print()
        print("  ▼ CellRichText のラン構造:")
        for i, run in enumerate(val):
            if isinstance(run, TextBlock):
                strike = bool(run.font and run.font.strike)
                bold   = bool(run.font and run.font.bold)
                print(f"    [{i}] TextBlock: text={repr(run.text)}, strike={strike}, bold={bold}")
            else:
                print(f"    [{i}] 文字列ラン: text={repr(run)}")
    else:
        font = target_cell.font
        print(f"セルフォント strike: {font.strike if font else 'None'}")

    print()
    print("=" * 60)
    print("【チェック5】cell_to_markdown() の出力")
    print("=" * 60)

    from excel_reader import cell_to_markdown
    md = cell_to_markdown(target_cell)
    print(f"戻り値: {repr(md)}")
    print()
    print("実際の内容（改行を視覚化）:")
    print(md)

    print()
    print("=" * 60)
    print("【チェック6】rich_text 設定の確認")
    print("=" * 60)

    if not rich_text_cfg:
        print("rich_text: false（または未設定）→ 取り消し線は Markdown に反映されません")
        print("→ config.yaml の issue_mapping に 'rich_text: true' を追加してください")
    else:
        print("rich_text: true ✓")

    desc_format = source.get("issue_mapping", {}).get("description_format", "template")
    print(f"description_format: {desc_format}")
    if desc_format == "template":
        tmpl = source.get("issue_mapping", {}).get("description_template", "")
        if "{{auto}}" in tmpl:
            print("→ {{auto}} プレースホルダーあり: 取り消し線が反映されます ✓")
        elif f"{{{{{col_name}}}}}" in tmpl:
            print(f"→ {{{{{{col_name}}}}}} はプレーンテキスト展開のみ: 取り消し線は反映されません ✗")
            print("  取り消し線を反映するには description_format: 'auto' か {{auto}} を使ってください")
        else:
            print(f"→ テンプレートに '{col_name}' の参照が見つかりません")

    print()
    print("=" * 60)
    print("デバッグ完了")
    print("=" * 60)


if __name__ == "__main__":
    main()
