"""
元 Excel を変更しないことのテスト
=================================
このツールは読み込んだ Excel ファイルを一切書き換えない。

openpyxl でファイルを開き直して保存すると、data_only=True で読んでいる
関係で数式が失われるほか、Excel 側で作成したグラフ・ピボットテーブル・
画像などが失われる可能性がある。元ファイルは業務で使われている実データの
ため、書き込みは行わない方針とする。

将来 issueKey の書き戻しなどを実装しようとした場合、このテストが落ちる。
"""

import hashlib
import os

import pytest
from openpyxl import Workbook, load_workbook

import excel_to_backlog as etb
from mapper import BacklogMaster
from run_log import RunLog


@pytest.fixture(autouse=True)
def _no_sleep(monkeypatch):
    monkeypatch.setattr(etb.time, "sleep", lambda _: None)


@pytest.fixture
def excel(tmp_path):
    """数式とデータを含むブック（数式の消失を検知するため）。"""
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"], ws["C1"] = "件名", "Backlog番号", "件数"
    for i, name in enumerate(["課題A", "課題B"], start=2):
        ws[f"A{i}"] = name
    ws["C2"] = "=COUNTA(A2:A3)"
    path = tmp_path / "source.xlsx"
    wb.save(path)
    return path


def digest(path):
    return hashlib.sha256(path.read_bytes()).hexdigest()


class FakeClient:
    def __init__(self):
        self.created = 0

    def create_issue(self, params):
        self.created += 1
        return {"issueKey": f"PROJ-{self.created}", "summary": params["summary"]}

    def get_issue(self, issue_key):
        return None


def source_cfg(excel_path, **extra):
    cfg = {
        "name": "テスト",
        "excel": {"path": str(excel_path)},
        "issue_mapping": {"issue_type": "タスク", "priority": "中", "summary_col": "件名"},
    }
    cfg.update(extra)
    return cfg


@pytest.fixture
def master():
    return BacklogMaster(project_id=1, issue_type_map={"タスク": 1}, priority_map={"中": 3})


class TestSourceExcelUntouched:
    def test_課題を作成しても元ファイルは変わらない(self, excel, master):
        before, before_mtime = digest(excel), os.path.getmtime(excel)

        counts = etb.process_source(
            source_cfg(excel), FakeClient(), master, dry_run=False
        )

        assert counts["created"] == 2
        assert digest(excel) == before
        assert os.path.getmtime(excel) == before_mtime

    def test_key_col_を使う設定でも元ファイルは変わらない(self, excel, master):
        """issueKey を書き戻す動機が最も強い設定でも書き込まないこと。"""
        before = digest(excel)

        etb.process_source(
            source_cfg(excel, upsert={"enabled": True, "key_col": "Backlog番号"}),
            FakeClient(), master, dry_run=False,
        )

        assert digest(excel) == before

    def test_実行ログを出力しても元ファイルは変わらない(self, excel, master, tmp_path):
        before = digest(excel)

        with RunLog(tmp_path / "run.csv") as log:
            etb.process_source(
                source_cfg(excel), FakeClient(), master, dry_run=False, run_log=log
            )

        assert digest(excel) == before
        assert (tmp_path / "run.csv").exists()

    def test_ドライランでも元ファイルは変わらない(self, excel, master):
        before = digest(excel)
        etb.process_source(source_cfg(excel), FakeClient(), master, dry_run=True)
        assert digest(excel) == before

    def test_プレビュー生成でも元ファイルは変わらない(self, excel, master, tmp_path):
        before = digest(excel)

        etb.generate_preview_for_source(
            source_cfg(excel), master, {}, tmp_path / "preview.md", "2026-01-01"
        )

        assert digest(excel) == before
        assert (tmp_path / "preview.md").exists()

    def test_数式が保持されている(self, excel, master):
        """
        openpyxl は data_only=True で読み込むため、保存し直すと数式が消える。
        保存していないことを数式の生存で確認する。
        """
        etb.process_source(source_cfg(excel), FakeClient(), master, dry_run=False)

        assert load_workbook(excel).active["C2"].value == "=COUNTA(A2:A3)"
