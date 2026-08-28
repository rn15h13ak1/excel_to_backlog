"""
件名インデックスのテスト
========================
行ごとの検索（N+1）を一括取得＋辞書引きに置き換えた部分。
API 呼び出し回数が行数に比例しないことを固定する。
"""

import pytest
from openpyxl import Workbook

import excel_to_backlog as etb
from mapper import BacklogMaster
from summary_index import SummaryIndex


class FakeClient:
    """get_issues を1回だけ許すクライアント。"""

    def __init__(self, issues=()):
        self.issues = list(issues)
        self.get_issues_calls = 0
        self.created = []

    def get_issues(self, project_id, params=None):
        self.get_issues_calls += 1
        return self.issues

    def create_issue(self, params):
        key = f"PROJ-{100 + len(self.created)}"
        self.created.append(params["summary"])
        return {"issueKey": key, "summary": params["summary"]}

    def update_issue(self, key, params):
        return {"issueKey": key}


def issue(key, summary):
    return {"issueKey": key, "summary": summary}


class TestSummaryIndex:
    def test_件名から_issueKey_を引ける(self):
        client = FakeClient([issue("PROJ-1", "ログイン不具合")])
        index = SummaryIndex(client, project_id=1)
        assert index.find("ログイン不具合") == "PROJ-1"

    def test_存在しない件名は_None(self):
        index = SummaryIndex(FakeClient([issue("PROJ-1", "A")]), project_id=1)
        assert index.find("存在しない") is None

    def test_参照するまで_API_を呼ばない(self):
        """match_summary を使わない設定で無駄な全件取得をしないこと。"""
        client = FakeClient([issue("PROJ-1", "A")])
        SummaryIndex(client, project_id=1)
        assert client.get_issues_calls == 0

    def test_何度引いても取得は1回だけ(self):
        client = FakeClient([issue("PROJ-1", "A"), issue("PROJ-2", "B")])
        index = SummaryIndex(client, project_id=1)
        for _ in range(10):
            index.find("A")
            index.find("B")
        assert client.get_issues_calls == 1

    def test_件名は正規化して照合される(self):
        """改行やタブを含む既存件名とも一致すること。"""
        client = FakeClient([issue("PROJ-1", "ログイン\n不具合")])
        index = SummaryIndex(client, project_id=1)
        assert index.find("ログイン不具合") == "PROJ-1"

    def test_件名が重複する場合は最初の1件(self, capsys):
        client = FakeClient([issue("PROJ-1", "重複"), issue("PROJ-2", "重複")])
        index = SummaryIndex(client, project_id=1)
        assert index.find("重複") == "PROJ-1"
        assert "重複している課題" in capsys.readouterr().err

    def test_空の件名は索引に入れない(self):
        client = FakeClient([issue("PROJ-1", ""), issue("PROJ-2", "A")])
        index = SummaryIndex(client, project_id=1)
        assert index.find("") is None
        assert index.find("A") == "PROJ-2"

    def test_作成した課題を索引に追加できる(self):
        client = FakeClient([issue("PROJ-1", "既存")])
        index = SummaryIndex(client, project_id=1)
        index.find("既存")            # 構築させる
        index.add("新規", "PROJ-9")
        assert index.find("新規") == "PROJ-9"

    def test_構築前の_add_は取得を先送りする(self):
        """次の find() の全件取得にその課題も含まれるため、追加は不要。"""
        client = FakeClient([issue("PROJ-1", "A")])
        index = SummaryIndex(client, project_id=1)
        index.add("新規", "PROJ-9")
        assert client.get_issues_calls == 0


class TestNoNPlusOne:
    """行数が増えても検索リクエストが増えないこと。"""

    @pytest.fixture(autouse=True)
    def _no_sleep(self, monkeypatch):
        monkeypatch.setattr(etb.time, "sleep", lambda _: None)

    @pytest.fixture
    def excel(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "件名"
        for i in range(2, 52):          # 50 行
            ws[f"A{i}"] = f"課題{i - 1}"
        path = tmp_path / "many.xlsx"
        wb.save(path)
        return str(path)

    def test_50行でも全件取得は1回(self, excel):
        client = FakeClient([issue("PROJ-1", "課題1")])
        index = SummaryIndex(client, project_id=1)
        source_cfg = {
            "name": "多数行",
            "excel": {"path": excel},
            "issue_mapping": {"issue_type": "タスク", "priority": "中", "summary_col": "件名"},
            "upsert": {"enabled": True, "match_summary": True},
        }
        counts = etb.process_source(
            source_cfg, client,
            BacklogMaster(project_id=1, issue_type_map={"タスク": 1}, priority_map={"中": 3}),
            dry_run=False, summary_index=index,
        )

        assert client.get_issues_calls == 1     # 50 行でも 1 回
        assert counts["updated"] == 1           # 既存の「課題1」
        assert counts["created"] == 49

    def test_同じ件名の行が2つあっても重複作成しない(self, tmp_path):
        """索引は起動時のスナップショットのため、作成分を追加していないと重複する。"""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "件名"
        ws["A2"], ws["A3"] = "同じ件名", "同じ件名"
        path = tmp_path / "dup.xlsx"
        wb.save(path)

        client = FakeClient([])
        index = SummaryIndex(client, project_id=1)
        counts = etb.process_source(
            {
                "name": "重複",
                "excel": {"path": str(path)},
                "issue_mapping": {"issue_type": "タスク", "priority": "中", "summary_col": "件名"},
                "upsert": {"enabled": True, "match_summary": True},
            },
            client,
            BacklogMaster(project_id=1, issue_type_map={"タスク": 1}, priority_map={"中": 3}),
            dry_run=False, summary_index=index,
        )

        assert counts["created"] == 1
        assert counts["updated"] == 1
        assert client.created == ["同じ件名"]
