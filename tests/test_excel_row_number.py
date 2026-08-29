"""
Excel 行番号のテスト
====================
表示・ログ・警告の行番号は、フィルタ後の連番ではなくシート上の行番号を使う。
連番ではシートの何行目か辿れず、失敗した行を特定できない。
"""

import csv

import excel_to_backlog as etb
from conftest import FakeBacklog
from excel_reader import ExcelReader
from run_log import RunLog

KEY = ExcelReader.ROW_NUMBER_KEY


class TestReaderAttachesRowNumber:
    def test_行データに行番号が入る(self, make_excel):
        path = make_excel(["件名"], [["課題A"], ["課題B"]])
        _, rows = ExcelReader({"path": str(path)}).read()

        assert [r[KEY] for r in rows] == ["2", "3"]      # ヘッダーが1行目

    def test_ヘッダー位置がずれても実際の行番号になる(self, make_excel, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A3"] = "件名"
        ws["A4"], ws["A5"] = "課題A", "課題B"
        path = tmp_path / "offset.xlsx"
        wb.save(path)

        _, rows = ExcelReader(
            {"path": str(path), "header_start_row": 3, "data_start_row": 4}
        ).read()

        assert [r[KEY] for r in rows] == ["4", "5"]

    def test_空行を飛ばしても行番号は詰まらない(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "件名"
        ws["A2"], ws["A4"] = "課題A", "課題B"      # 3行目は空
        path = tmp_path / "gap.xlsx"
        wb.save(path)

        _, rows = ExcelReader({"path": str(path)}).read()

        assert [r[KEY] for r in rows] == ["2", "4"]

    def test_書式付き行にも同じ行番号が入る(self, make_excel):
        path = make_excel(["件名"], [["課題A"], ["課題B"]])
        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert [r[KEY] for r in plain] == [r[KEY] for r in formatted] == ["2", "3"]


class TestRowNumberInOutput:
    def test_フィルタ後も元の行番号が表示される(self, source_cfg, master, capsys):
        cfg = source_cfg(
            ["件名", "状態"],
            [["課題A", "対応要"], ["除外", "完了"], ["課題B", "対応要"]],
            filters=[{"col_name": "状態", "value": "対応要"}],
        )

        etb.process_source(cfg, FakeBacklog(), master, dry_run=False)

        out = capsys.readouterr().out
        assert "[2] ✅ 作成" in out      # 2行目
        assert "[4] ✅ 作成" in out      # 4行目（3行目は除外）
        assert "[3]" not in out

    def test_実行ログにも元の行番号が入る(self, source_cfg, master, tmp_path):
        cfg = source_cfg(
            ["件名", "状態"],
            [["課題A", "対応要"], ["除外", "完了"], ["課題B", "対応要"]],
            filters=[{"col_name": "状態", "value": "対応要"}],
        )
        log_path = tmp_path / "run.csv"

        with RunLog(log_path) as log:
            etb.process_source(cfg, FakeBacklog(), master, dry_run=False, run_log=log)

        with open(log_path, encoding="utf-8-sig", newline="") as f:
            assert [r["row"] for r in csv.DictReader(f)] == ["2", "4"]

    def test_ドライランでも同じ行番号(self, source_cfg, master, capsys):
        cfg = source_cfg(["件名"], [["課題A"], ["課題B"]])
        etb.process_source(cfg, FakeBacklog(), master, dry_run=True)

        out = capsys.readouterr().out
        assert "[2]" in out and "[3]" in out


class TestRowNumberIsNotAColumn:
    def test_本文の自動生成に行番号が出ない(self, source_cfg, master):
        """メタキーは Excel の列ではないので description に含めない。"""
        cfg = source_cfg(
            ["件名", "概要"], [["課題A", "本文"]],
            issue_mapping={"description_format": "auto"},
        )
        backlog = FakeBacklog()
        etb.process_source(cfg, backlog, master, dry_run=False)

        # FakeBacklog は summary しか保持しないため、mapper を直接使って確認する
        from mapper import IssueMapper
        m = IssueMapper(
            {"issue_type": "タスク", "priority": "中", "summary_col": "件名",
             "description_format": "auto"},
            master, headers=["件名", "概要"],
        )
        params = m.map_row({"件名": "課題A", "概要": "本文", KEY: "2"})
        assert KEY not in params["description"]

    def test_テンプレートから参照できる(self, master):
        from mapper import IssueMapper
        m = IssueMapper(
            {"issue_type": "タスク", "priority": "中",
             "summary_template": "行{{_excel_row}}: {{件名}}"},
            master, headers=["件名"],
        )
        params = m.map_row({"件名": "課題A", KEY: "7"})
        assert params["summary"] == "行7: 課題A"

    def test_列名検証で既知として扱われる(self):
        from excel_to_backlog import validate_column_references
        cfg = {"issue_mapping": {"summary_template": "行{{_excel_row}}: {{件名}}"}}
        assert validate_column_references(cfg, ["件名"]) == []
