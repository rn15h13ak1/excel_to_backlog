"""
Strict Open XML の判定テスト
============================
openpyxl は Strict Open XML に対応しておらず、そのまま読み込むと共有文字列
テーブルが空になり "list index out of range" という原因と無関係なエラーになる。
事前に判定して対処を示せることを固定する。
"""

import zipfile

import pytest
from openpyxl import Workbook

from excel_reader import ExcelReader


@pytest.fixture
def normal_xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["A2"] = "件名", "テスト"
    path = tmp_path / "normal.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def strict_xlsx(normal_xlsx, tmp_path):
    """workbook.xml の名前空間を Strict に差し替えた xlsx を作る。"""
    path = tmp_path / "strict.xlsx"
    with zipfile.ZipFile(normal_xlsx) as src, zipfile.ZipFile(path, "w") as dst:
        for item in src.namelist():
            data = src.read(item)
            if item.endswith("workbook.xml") and "_rels" not in item:
                data = data.replace(
                    ExcelReader.TRANSITIONAL_NS.encode(),
                    ExcelReader.STRICT_NS.encode(),
                )
            dst.writestr(item, data)
    return path


class TestDetectOoxmlVariant:
    def test_通常の_xlsx_は_transitional(self, normal_xlsx):
        assert ExcelReader.detect_ooxml_variant(normal_xlsx) == "transitional"

    def test_Strict_形式を判定できる(self, strict_xlsx):
        assert ExcelReader.detect_ooxml_variant(strict_xlsx) == "strict"

    def test_zip_でないファイルは_unknown(self, tmp_path):
        path = tmp_path / "not_a_zip.xlsx"
        path.write_text("これは xlsx ではありません", encoding="utf-8")
        assert ExcelReader.detect_ooxml_variant(path) == "unknown"

    def test_workbook_xml_がなければ_unknown(self, tmp_path):
        path = tmp_path / "empty.xlsx"
        with zipfile.ZipFile(path, "w") as z:
            z.writestr("dummy.txt", "x")
        assert ExcelReader.detect_ooxml_variant(path) == "unknown"


class TestStrictRejection:
    def test_読み込み時に対処を示して停止する(self, strict_xlsx):
        reader = ExcelReader({"path": str(strict_xlsx)})
        with pytest.raises(ValueError) as exc:
            reader.read()

        message = str(exc.value)
        assert "Strict Open XML" in message
        assert "名前を付けて保存" in message

    def test_通常の_xlsx_は読み込める(self, normal_xlsx):
        headers, rows = ExcelReader({"path": str(normal_xlsx)}).read()
        assert headers == ["件名"]
        assert rows[0]["件名"] == "テスト"
