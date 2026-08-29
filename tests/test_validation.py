"""
列名参照の事前検証のテスト
==========================
列名が1つでも一致しないと、フィルター条件が消えて全行が対象になる・
プレースホルダーが未展開のまま件名になる、といった無言の誤動作が起きる。
実行前に必ず検出できることを固定する。
"""

import pytest

from excel_to_backlog import (
    collect_referenced_columns,
    new_counts,
    process_source,
    validate_column_references,
)
from mapper import BacklogMaster

HEADERS = ["項番", "枝番", "件名", "ステータス", "担当者", "期限"]


def unknown_columns(source_cfg, headers=HEADERS):
    """検出された未知の列名だけを取り出す。"""
    known = set(headers) | {"_source_name", "_excel_path", "_excel_sheet"}
    return [col for _, col in collect_referenced_columns(source_cfg) if col not in known]


class TestCollectReferencedColumns:
    def test_フィルターの列名を集める(self):
        cfg = {"filters": [{"col_name": "ステータス", "value": "x"}]}
        assert ("filters[0].col_name", "ステータス") in collect_referenced_columns(cfg)

    def test_filter_groups_の列名も集める(self):
        cfg = {"filter_groups": [{"filters": [{"col_name": "項番", "value": "1"}]}]}
        paths = [p for p, _ in collect_referenced_columns(cfg)]
        assert "filter_groups[0].filters[0].col_name" in paths

    def test_summary_template_があれば_summary_col_は見ない(self):
        cfg = {"issue_mapping": {"summary_template": "{{件名}}", "summary_col": "使わない列"}}
        cols = [c for _, c in collect_referenced_columns(cfg)]
        assert "件名" in cols
        assert "使わない列" not in cols

    def test_auto_モードでは_description_template_を見ない(self):
        cfg = {"issue_mapping": {
            "summary_col": "件名",
            "description_format": "auto",
            "description_template": "{{古い列}}",
        }}
        assert "古い列" not in [c for _, c in collect_referenced_columns(cfg)]

    def test_メタキーは既知として扱う(self):
        cfg = {"issue_mapping": {"summary_template": "【{{_source_name}}】{{件名}}"}}
        assert unknown_columns(cfg) == []


class TestValidateColumnReferences:
    def test_すべて解決できればエラーなし(self):
        cfg = {
            "filters": [{"col_name": "ステータス", "value": "対応要"}],
            "issue_mapping": {
                "summary_template": "項番{{項番}}{{#枝番}}-{{枝番}}{{/枝番}} {{件名}}",
                "description_template": "担当: {{担当者}}\n{{auto}}",
                "due_date_col": "期限",
                "assignee_col": "担当者",
                "status_col": "ステータス",
                "required_cols": ["件名"],
            },
        }
        assert validate_column_references(cfg, HEADERS) == []

    def test_末尾の空白は不一致として検出される(self):
        """
        filter_rows() は設定値を strip せずに突き合わせるため、
        末尾に空白があると条件が消えて全行が通ってしまう。
        """
        cfg = {"filters": [{"col_name": "ステータス ", "value": "対応要"}]}
        errors = validate_column_references(cfg, HEADERS)
        assert errors, "末尾スペースを検出できていない"
        assert any("前後の空白" in line for line in errors)

    def test_テンプレートの未知の列を検出する(self):
        cfg = {"issue_mapping": {"summary_template": "【{{分類}}】{{件名}}"}}
        errors = validate_column_references(cfg, HEADERS)
        assert any("分類" in line for line in errors)

    def test_カスタム属性の列名を検出する(self):
        cfg = {"issue_mapping": {
            "summary_col": "件名",
            "custom_fields": [{"field_name": "カテゴリ", "col_name": "区分"}],
        }}
        assert any("区分" in line for line in validate_column_references(cfg, HEADERS))

    def test_upsert_の_key_col_を検出する(self):
        cfg = {"issue_mapping": {"summary_col": "件名"},
               "upsert": {"enabled": True, "key_col": "Backlog番号"}}
        assert any("Backlog番号" in line for line in validate_column_references(cfg, HEADERS))

    def test_複数の誤りをまとめて報告する(self):
        cfg = {
            "filters": [{"col_name": "無い列A", "value": "x"}],
            "issue_mapping": {"summary_template": "{{無い列B}}", "due_date_col": "無い列C"},
        }
        errors = validate_column_references(cfg, HEADERS)
        assert "3 件" in errors[0]

    def test_テンプレートの空白は許容される(self):
        """{{ 件名 }} は _render_template() が strip するため一致とみなす。"""
        cfg = {"issue_mapping": {"summary_template": "{{ 件名 }}"}}
        assert validate_column_references(cfg, HEADERS) == []


class TestProcessSourceAborts:
    """検証に失敗した場合、API を1度も呼ばずに中止すること。"""

    @pytest.fixture
    def excel(self, tmp_path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for col, header in zip("ABC", ["項番", "件名", "ステータス"]):
            ws[f"{col}1"] = header
        for i, (name, status) in enumerate(
            [("ログイン不具合", "対応要"), ("表示崩れ", "完了"), ("速度改善", "完了")], start=2
        ):
            ws[f"A{i}"], ws[f"B{i}"], ws[f"C{i}"] = i - 1, name, status
        path = tmp_path / "src.xlsx"
        wb.save(path)
        return str(path)

    def test_列名が一致しなければ_API_を呼ばずにエラーを返す(self, excel):
        class ExplodingClient:
            def __getattr__(self, name):
                raise AssertionError(f"API が呼ばれた: {name}()")

        source_cfg = {
            "name": "テスト",
            "excel": {"path": excel},
            "filters": [{"col_name": "ステータス ", "value": "対応要"}],
            "issue_mapping": {"issue_type": "タスク", "summary_col": "件名"},
        }
        counts = process_source(
            source_cfg, ExplodingClient(), BacklogMaster(project_id=1), dry_run=False
        )

        expected = new_counts()
        expected["error"] = 1
        assert counts == expected


class TestNotValidatedWhenUnused:
    """
    実行時に読まれない設定項目は検証しない。
    「設定に書かれているか」ではなく「実際に読まれるか」で判断する。
    """

    def test_filter_groups_があれば_filters_は検証しない(self):
        """apply_filters() は filter_groups を優先し filters を無視する。"""
        cfg = {
            "filters": [{"col_name": "廃止した列", "value": "x"}],
            "filter_groups": [{"filters": [{"col_name": "項番", "value": "1"}]}],
            "issue_mapping": {"summary_col": "件名"},
        }
        assert validate_column_references(cfg, HEADERS) == []

    def test_filter_groups_がなければ_filters_を検証する(self):
        cfg = {
            "filters": [{"col_name": "無い列", "value": "x"}],
            "issue_mapping": {"summary_col": "件名"},
        }
        assert validate_column_references(cfg, HEADERS)

    def test_upsert_無効なら_key_col_は検証しない(self):
        """書き戻し列を用意する前に設定だけ書いておける。"""
        cfg = {
            "issue_mapping": {"summary_col": "件名"},
            "upsert": {"enabled": False, "key_col": "Backlog番号"},
        }
        assert validate_column_references(cfg, HEADERS) == []

    def test_upsert_有効なら_key_col_を検証する(self):
        cfg = {
            "issue_mapping": {"summary_col": "件名"},
            "upsert": {"enabled": True, "key_col": "Backlog番号"},
        }
        assert validate_column_references(cfg, HEADERS)

    def test_auto_を使わない_template_では_description_cols_を検証しない(self):
        """_render_auto() が動かないため description_cols は読まれない。"""
        cfg = {"issue_mapping": {
            "summary_col": "件名",
            "description_format": "template",
            "description_template": "{{件名}}",
            "description_cols": ["廃止した列"],
        }}
        assert validate_column_references(cfg, HEADERS) == []

    def test_auto_モードなら_description_cols_を検証する(self):
        cfg = {"issue_mapping": {
            "summary_col": "件名",
            "description_format": "auto",
            "description_cols": ["無い列"],
        }}
        assert validate_column_references(cfg, HEADERS)

    def test_テンプレートに_auto_があれば_description_cols_を検証する(self):
        cfg = {"issue_mapping": {
            "summary_col": "件名",
            "description_template": "見出し\n{{auto}}",
            "description_cols": ["無い列"],
        }}
        assert validate_column_references(cfg, HEADERS)

    def test_空白付きの_auto_も認識する(self):
        cfg = {"issue_mapping": {
            "summary_col": "件名",
            "description_template": "{{ auto }}",
            "description_cols": ["無い列"],
        }}
        assert validate_column_references(cfg, HEADERS)
