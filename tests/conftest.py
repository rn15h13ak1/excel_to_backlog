"""
テスト共通のフィクスチャ
========================
各テストファイルに散らばっていた FakeClient と Excel 生成処理をまとめる。

FakeBacklog は課題の状態を保持するため、「作成 → 件名を変更 → 再実行」
のような複数回の実行にまたがる振る舞いを検証できる。
"""

import re
import zipfile

import pytest
from openpyxl import Workbook

import backlog_client
import excel_to_backlog as etb
from mapper import BacklogMaster


@pytest.fixture(autouse=True)
def no_sleep(monkeypatch):
    """レート制限用の待機とリトライ待機を無効化する（全テスト共通）。"""
    monkeypatch.setattr(etb.time, "sleep", lambda _: None)
    monkeypatch.setattr(backlog_client.time, "sleep", lambda _: None)


# ------------------------------------------------------------------
# Backlog のふるまいを模したクライアント
# ------------------------------------------------------------------

class FakeBacklog:
    """
    課題の状態を保持する BacklogClient の代替。

    Parameters
    ----------
    issues : dict[str, str] | None
        あらかじめ存在する課題 {issueKey: 件名}
    fail_create_at : int | None
        N 回目の create_issue で BacklogAPIError を送出する
    fail_create_matching : str | None
        件名にこの文字列を含む課題の作成で BacklogAPIError を送出する
    fail_status_update : bool
        ステータス変更（statusId のみの更新）で失敗させる
    no_change_on_update : bool
        更新時に常に BacklogNoChangeError を送出する
    """

    def __init__(
        self,
        issues=None,
        *,
        fail_create_at=None,
        fail_create_matching=None,
        fail_status_update=False,
        no_change_on_update=False,
    ):
        self.issues = dict(issues or {})
        self.fail_create_at = fail_create_at
        self.fail_create_matching = fail_create_matching
        self.fail_status_update = fail_status_update
        self.no_change_on_update = no_change_on_update

        self.create_calls = 0
        self.get_issues_calls = 0
        self.updates = []          # [(issueKey, params), ...]
        self._next_id = len(self.issues)

    # ---- マスターデータ ----

    def get_project(self, project_key):
        return {"id": 42, "projectKey": project_key}

    def get_issue_types(self, project_key):
        return [{"name": "タスク", "id": 1}, {"name": "バグ", "id": 2}]

    def get_priorities(self):
        return [{"name": "高", "id": 2}, {"name": "中", "id": 3}]

    def get_project_users(self, project_key):
        return [{"name": "山田太郎", "id": 10, "userId": "yamada"}]

    def get_custom_fields(self, project_key):
        return []

    def get_statuses(self, project_key):
        return [{"name": "未対応", "id": 1}, {"name": "完了", "id": 4}]

    # ---- 課題 ----

    def get_issue(self, issue_id_or_key):
        key = str(issue_id_or_key)
        if key not in self.issues:
            return None
        return {"issueKey": key, "summary": self.issues[key]}

    def get_issues(self, project_id, params=None):
        self.get_issues_calls += 1
        return [{"issueKey": k, "summary": s} for k, s in self.issues.items()]

    def create_issue(self, params):
        self.create_calls += 1
        summary = params["summary"]

        if self.fail_create_at == self.create_calls:
            raise backlog_client.BacklogAPIError("作成に失敗しました", status=400)
        if self.fail_create_matching and self.fail_create_matching in summary:
            raise backlog_client.BacklogAPIError(
                "作成に失敗しました（件名が不正）", status=400
            )

        self._next_id += 1
        key = f"DEMO-{self._next_id}"
        self.issues[key] = summary
        return {"issueKey": key, "summary": summary}

    def update_issue(self, issue_id_or_key, params):
        key = str(issue_id_or_key)
        self.updates.append((key, params))

        if self.fail_status_update and set(params) == {"statusId"}:
            raise backlog_client.BacklogAPIError(
                "ステータスの変更が許可されていません", status=400
            )
        if self.no_change_on_update:
            raise backlog_client.BacklogNoChangeError("変更されていません（code=7）")

        if "summary" in params:
            self.issues[key] = params["summary"]
        return {"issueKey": key, "summary": self.issues.get(key, "")}


class ExplodingBacklog:
    """API が呼ばれたら失敗するクライアント（呼ばれないことの検証用）。"""

    def __getattr__(self, name):
        raise AssertionError(f"API が呼ばれた: {name}()")


# ------------------------------------------------------------------
# フィクスチャ
# ------------------------------------------------------------------

@pytest.fixture
def master():
    """FakeBacklog のマスターデータに対応する BacklogMaster。"""
    return BacklogMaster(
        project_id=42,
        issue_type_map={"タスク": 1, "バグ": 2},
        priority_map={"高": 2, "中": 3},
        user_map={"山田太郎": 10, "yamada": 10},
        status_map={"未対応": 1, "完了": 4},
    )


@pytest.fixture
def make_excel(tmp_path):
    """
    ヘッダーと行データから xlsx を作るファクトリ。

        path = make_excel(["件名", "期限"], [["課題A", "2025/01/05"]])
    """
    counter = {"n": 0}

    def _make(headers, rows, name=None):
        counter["n"] += 1
        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        for r, row in enumerate(rows, start=2):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)
        path = tmp_path / (name or f"source{counter['n']}.xlsx")
        wb.save(path)
        return path

    return _make


@pytest.fixture
def source_cfg(make_excel):
    """
    最小構成の sources[i] を作るファクトリ。

        cfg = source_cfg(["件名"], [["課題A"]], upsert={"enabled": True})
    """
    def _make(headers, rows, *, name="テスト", issue_mapping=None, **extra):
        mapping = {"issue_type": "タスク", "priority": "中", "summary_col": "件名"}
        mapping.update(issue_mapping or {})
        cfg = {
            "name": name,
            "excel": {"path": str(make_excel(headers, rows))},
            "issue_mapping": mapping,
        }
        cfg.update(extra)
        return cfg

    return _make


# ------------------------------------------------------------------
# リッチテキスト（取り消し線）を含む Excel の生成
# ------------------------------------------------------------------

def _runs_xml(runs) -> str:
    """
    [(テキスト, 取り消し線 or None), ...] を <is> 要素の XML にする。

    None を渡すと <rPr> を持たないランになる。openpyxl はこれを素の str として
    返し、<rPr> を持つランを TextBlock として返す。Excel が「セル全体に
    取り消し線 → 一部を解除」した際の保存形式を再現できる。
    """
    parts = []
    for text, struck in runs:
        rpr = "" if struck is None else (
            "<rPr><strike/></rPr>" if struck else '<rPr><strike val="false"/></rPr>'
        )
        parts.append(f"<r>{rpr}<t xml:space=\"preserve\">{text}</t></r>")
    return "<is>" + "".join(parts) + "</is>"


@pytest.fixture
def make_rich_excel(tmp_path):
    """
    リッチテキストのセルを含む xlsx を作るファクトリ。

        path = make_rich_excel(
            ["件名", "対応内容"],
            [["課題A", [("残る", False), ("消える", True)]]],
        )

    セルの値には文字列か、ラン定義のリストを渡す。

    openpyxl 3.1 は CellRichText を含むブックを保存できないため、
    いったん目印を書いたブックを保存し、シートの XML を差し替えて作る。
    openpyxl は文字列セルを t="inlineStr" で書くため、<is> の中身を
    ランに置き換えればよい（sharedStrings は生成されない）。
    """
    counter = {"n": 0}

    def _make(headers, rows, name=None):
        counter["n"] += 1
        wb = Workbook()
        ws = wb.active
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)

        markers = {}
        for r, row in enumerate(rows, start=2):
            for c, value in enumerate(row, start=1):
                if isinstance(value, list):
                    marker = f"@@RICH{len(markers)}@@"
                    markers[marker] = _runs_xml(value)
                    ws.cell(row=r, column=c, value=marker)
                else:
                    ws.cell(row=r, column=c, value=value)

        base = tmp_path / f"_base{counter['n']}.xlsx"
        wb.save(base)

        path = tmp_path / (name or f"rich{counter['n']}.xlsx")
        with zipfile.ZipFile(base) as src, zipfile.ZipFile(path, "w") as dst:
            for item in src.namelist():
                data = src.read(item)
                if item == "xl/worksheets/sheet1.xml":
                    xml = data.decode()
                    for marker, runs_xml in markers.items():
                        xml = re.sub(f"<is><t[^>]*>{marker}</t></is>", runs_xml, xml)
                    data = xml.encode()
                dst.writestr(item, data)
        base.unlink()
        return path

    return _make
