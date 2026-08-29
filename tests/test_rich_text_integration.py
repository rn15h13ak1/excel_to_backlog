"""
rich_text の結合テスト
======================
ファイルを読んで平文行と書式付き行を対応づける経路（read_with_format /
_build_rows_dual）を通す。cell_to_markdown 単体のテストはあるが、この
結合部分は一度も実行されていなかった。

取り消し線の継承ロジックは追加・削除・再追加と往復した箇所であり、
平文行と書式付き行の対応づけは id(dict) に依存している。対応が崩れても
例外は出ず、本文だけが静かに元テキストに戻るため、テストで固定する。
"""

import excel_to_backlog as etb
from conftest import FakeBacklog
from excel_reader import ExcelReader

KEY = ExcelReader.ROW_NUMBER_KEY


class TestReadWithFormat:
    def test_平文行と書式付き行が同じ件数になる(self, make_rich_excel):
        path = make_rich_excel(
            ["件名", "対応内容"],
            [["課題A", [("残る", False), ("消える", True)]],
             ["課題B", "書式なし"]],
        )

        headers, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert headers == ["件名", "対応内容"]
        assert len(plain) == len(formatted) == 2

    def test_平文行には取り消し線が入らない(self, make_rich_excel):
        """フィルタ・件名・担当者の解決には平文を使う。"""
        path = make_rich_excel(
            ["対応内容"], [[[("残る", False), ("消える", True)]]]
        )

        _, plain, _ = ExcelReader({"path": str(path)}).read_with_format()

        assert plain[0]["対応内容"] == "残る消える"

    def test_書式付き行に取り消し線が入る(self, make_rich_excel):
        path = make_rich_excel(
            ["対応内容"], [[[("残る", False), ("消える", True)]]]
        )

        _, _, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert formatted[0]["対応内容"] == "残る ~~消える~~"

    def test_両方に同じ行番号が入る(self, make_rich_excel):
        path = make_rich_excel(
            ["対応内容"], [["A"], [[("消える", True)]], ["C"]]
        )

        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert [r[KEY] for r in plain] == ["2", "3", "4"]
        assert [r[KEY] for r in formatted] == ["2", "3", "4"]

    def test_空行を飛ばしても両者の順序が一致する(self, make_rich_excel, tmp_path):
        """
        空行判定は平文で行う。書式付き側だけ判定基準が変わると件数がずれ、
        本文が別の行のものになる。
        """
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "対応内容"
        ws["A2"], ws["A4"] = "1行目", "3行目"      # 3行目は空
        path = tmp_path / "gap.xlsx"
        wb.save(path)

        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert len(plain) == len(formatted) == 2
        assert [r["対応内容"] for r in plain] == ["1行目", "3行目"]
        assert [r["対応内容"] for r in formatted] == ["1行目", "3行目"]

    def test_複数セルの書式が混在しても正しく対応する(self, make_rich_excel):
        path = make_rich_excel(
            ["件名", "備考"],
            [["課題A", [("旧仕様", True)]],
             ["課題B", "通常"],
             ["課題C", [("一部", False), ("削除", True)]]],
        )

        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert [r["備考"] for r in formatted] == ["~~旧仕様~~", "通常", "一部 ~~削除~~"]
        assert [r["備考"] for r in plain] == ["旧仕様", "通常", "一部削除"]


class TestRichTextInDescription:
    """rich_text: true のとき本文に取り消し線が反映されること。"""

    def _cfg(self, path, **mapping):
        base = {
            "issue_type": "タスク", "priority": "中", "summary_col": "件名",
            "rich_text": True, "description_format": "auto",
            "description_cols": ["対応内容"],
        }
        base.update(mapping)
        return {"name": "S", "excel": {"path": str(path)},
                "issue_mapping": base}

    def test_auto_モードの本文に反映される(self, make_rich_excel, master):
        path = make_rich_excel(
            ["件名", "対応内容"],
            [["課題A", [("残る", False), ("消える", True)]]],
        )
        backlog = FakeBacklog()

        etb.process_source(self._cfg(path), backlog, master, dry_run=False)

        # FakeBacklog は summary しか保持しないため plan_row で本文を確認する
        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()
        from mapper import IssueMapper
        mapper = IssueMapper(
            self._cfg(path)["issue_mapping"], master, headers=["件名", "対応内容"]
        )
        plan = etb.plan_row(
            plain[0], self._cfg(path), mapper, formatted_row=formatted[0]
        )
        assert "~~消える~~" in plan.params["description"]

    def test_テンプレートの_auto_にも反映される(self, make_rich_excel, master):
        path = make_rich_excel(
            ["件名", "対応内容"], [["課題A", [("消える", True)]]]
        )
        cfg = self._cfg(
            path, description_format="template",
            description_template="前書き\n{{auto}}",
        )
        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        from mapper import IssueMapper
        mapper = IssueMapper(cfg["issue_mapping"], master, headers=["件名", "対応内容"])
        plan = etb.plan_row(plain[0], cfg, mapper, formatted_row=formatted[0])

        assert "前書き" in plan.params["description"]
        assert "~~消える~~" in plan.params["description"]

    def test_列プレースホルダーは平文のまま(self, make_rich_excel, master):
        """{{列名}} は平文を使う（取り消し線を入れない）仕様。"""
        path = make_rich_excel(
            ["件名", "対応内容"], [["課題A", [("消える", True)]]]
        )
        cfg = self._cfg(
            path, description_format="template",
            description_template="内容: {{対応内容}}",
        )
        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        from mapper import IssueMapper
        mapper = IssueMapper(cfg["issue_mapping"], master, headers=["件名", "対応内容"])
        plan = etb.plan_row(plain[0], cfg, mapper, formatted_row=formatted[0])

        assert plan.params["description"] == "内容: 消える"

    def test_rich_text_無効なら書式を読まない(self, make_rich_excel, master):
        path = make_rich_excel(
            ["件名", "対応内容"], [["課題A", [("消える", True)]]]
        )
        cfg = self._cfg(path, rich_text=False)
        headers, rows = ExcelReader(cfg["excel"]).read()

        from mapper import IssueMapper
        mapper = IssueMapper(cfg["issue_mapping"], master, headers=headers)
        plan = etb.plan_row(rows[0], cfg, mapper)

        assert "~~" not in plan.params["description"]


class TestPlainFormattedAlignment:
    """
    process_source は id(dict) で平文行と書式付き行を対応づけている。
    フィルタで行が絞られても正しい本文が付くこと。
    """

    def test_フィルタ後も正しい書式が対応する(self, make_rich_excel, master):
        path = make_rich_excel(
            ["件名", "状態", "備考"],
            [["課題A", "対応要", [("Aの備考", True)]],
             ["除外",   "完了",   [("除外の備考", True)]],
             ["課題B", "対応要", [("Bの備考", True)]]],
        )
        cfg = {
            "name": "S", "excel": {"path": str(path)},
            "filters": [{"col_name": "状態", "value": "対応要"}],
            "issue_mapping": {
                "issue_type": "タスク", "priority": "中", "summary_col": "件名",
                "rich_text": True, "description_format": "auto",
                "description_cols": ["備考"],
            },
        }

        headers, plain, formatted = ExcelReader(cfg["excel"]).read_with_format()
        filtered = etb.apply_filters(plain, cfg, headers)
        index = {id(r): i for i, r in enumerate(plain)}

        from mapper import IssueMapper
        mapper = IssueMapper(cfg["issue_mapping"], master, headers=headers)
        bodies = [
            etb.plan_row(
                row, cfg, mapper, formatted_row=formatted[index[id(row)]]
            ).params["description"]
            for row in filtered
        ]

        assert "~~Aの備考~~" in bodies[0]
        assert "~~Bの備考~~" in bodies[1]
        assert all("除外" not in b for b in bodies)
