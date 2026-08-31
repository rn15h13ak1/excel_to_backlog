"""
使い勝手向けオプションのテスト
==============================
設定を書く前に必要な情報を確認する手段（--list-master / --show-columns）と、
初回に少数だけ試す手段（--limit）。
"""

import pytest

import excel_to_backlog as etb
from conftest import FakeBacklog
from mapper import BacklogMaster


class TestListMaster:
    @pytest.fixture
    def backlog(self):
        b = FakeBacklog()
        b.get_custom_fields = lambda k: [
            {"name": "カテゴリ", "id": 5, "typeId": 5,
             "items": [{"name": "設計", "id": 51}, {"name": "開発", "id": 52}]},
            {"name": "対応工数", "id": 6, "typeId": 3, "items": []},
        ]
        return b

    def test_設定に書ける名前がすべて出る(self, backlog, capsys):
        etb.print_master_data(BacklogMaster.build(backlog, "DEMO"))

        out = capsys.readouterr().out
        for expected in ["タスク", "バグ", "高", "中", "未対応", "完了", "山田太郎"]:
            assert expected in out

    def test_カスタム属性の型と選択肢が出る(self, backlog, capsys):
        """従来どこにも表示されず、正しい名前を知る手段がなかった。"""
        etb.print_master_data(BacklogMaster.build(backlog, "DEMO"))

        out = capsys.readouterr().out
        assert "カテゴリ  [単一リスト]" in out
        assert "設計 / 開発" in out
        assert "対応工数  [数値]" in out

    def test_担当者は表示名とログインIDをまとめて出す(self, backlog, capsys):
        etb.print_master_data(BacklogMaster.build(backlog, "DEMO"))
        assert "山田太郎 / yamada" in capsys.readouterr().out

    def test_カスタム属性が無い場合も分かるように出す(self, capsys):
        etb.print_master_data(BacklogMaster.build(FakeBacklog(), "DEMO"))
        assert "定義されていません" in capsys.readouterr().out

    def test_設定項目名が併記される(self, backlog, capsys):
        """どの設定キーに書けばよいかが分かること。"""
        etb.print_master_data(BacklogMaster.build(backlog, "DEMO"))

        out = capsys.readouterr().out
        assert "issue_mapping.issue_type" in out
        assert "custom_fields.field_name" in out


class TestShowColumns:
    def test_列記号と列名が出る(self, make_excel, capsys):
        path = make_excel(["項番", "件名"], [["1", "課題A"]])

        failures = etb.print_source_columns(
            [{"name": "S", "excel": {"path": str(path)}}]
        )

        out = capsys.readouterr().out
        assert failures == 0
        assert "A: 項番" in out
        assert "B: 件名" in out

    def test_実データの例が添えられる(self, make_excel, capsys):
        path = make_excel(["件名"], [["", ""], ["ログイン不具合"]])
        etb.print_source_columns([{"name": "S", "excel": {"path": str(path)}}])
        assert "例: ログイン不具合" in capsys.readouterr().out

    def test_複数行ヘッダーの結合結果が出る(self, tmp_path, capsys):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"], ws["A2"] = "大分類", "小分類"
        ws["A3"] = "値"
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        etb.print_source_columns([{
            "name": "S",
            "excel": {"path": str(path), "header_start_row": 1,
                      "header_end_row": 2, "data_start_row": 3},
        }])

        assert "大分類 / 小分類" in capsys.readouterr().out

    def test_同名の列は読み込まれない旨を出す(self, make_excel, capsys):
        path = make_excel(["備考", "備考"], [["A", "B"]])
        etb.print_source_columns([{"name": "S", "excel": {"path": str(path)}}])

        out = capsys.readouterr().out
        assert "B: 備考  ← 同名（本文には出力／列名指定は左端）" in out
        assert "(2)" not in out

    def test_読み込み失敗は件数として返る(self, tmp_path, capsys):
        failures = etb.print_source_columns(
            [{"name": "S", "excel": {"path": str(tmp_path / "ない.xlsx")}}]
        )
        assert failures == 1
        assert "読み込みに失敗" in capsys.readouterr().err

    def test_読込行数が出る(self, make_excel, capsys):
        path = make_excel(["件名"], [["A"], ["B"], ["C"]])
        etb.print_source_columns([{"name": "S", "excel": {"path": str(path)}}])
        assert "読込 3 行" in capsys.readouterr().out


class TestLimit:
    def test_先頭_N_行だけ処理する(self, source_cfg, master):
        cfg = source_cfg(["件名"], [[f"課題{i}"] for i in range(1, 11)])
        backlog = FakeBacklog()

        counts = etb.process_source(cfg, backlog, master, dry_run=False, limit=3)

        assert counts["created"] == 3
        assert backlog.create_calls == 3

    def test_フィルター後の先頭から数える(self, source_cfg, master):
        cfg = source_cfg(
            ["件名", "状態"],
            [["除外1", "完了"], ["対象1", "対応要"], ["除外2", "完了"], ["対象2", "対応要"]],
            filters=[{"col_name": "状態", "value": "対応要"}],
        )
        backlog = FakeBacklog()

        etb.process_source(cfg, backlog, master, dry_run=False, limit=1)

        assert backlog.issues == {"DEMO-1": "対象1"}

    def test_行数より大きい_limit_は影響しない(self, source_cfg, master):
        cfg = source_cfg(["件名"], [["A"], ["B"]])
        counts = etb.process_source(cfg, FakeBacklog(), master, dry_run=False, limit=99)
        assert counts["created"] == 2

    def test_制限した旨が表示される(self, source_cfg, master, capsys):
        cfg = source_cfg(["件名"], [[f"課題{i}"] for i in range(1, 6)])
        etb.process_source(cfg, FakeBacklog(), master, dry_run=False, limit=2)

        out = capsys.readouterr().out
        assert "--limit 2" in out
        assert "残り 3 行" in out

    def test_ドライランでも効く(self, source_cfg, master):
        cfg = source_cfg(["件名"], [[f"課題{i}"] for i in range(1, 11)])
        counts = etb.process_source(cfg, FakeBacklog(), master, dry_run=True, limit=2)
        assert counts["created"] == 2

    def test_未指定なら全行(self, source_cfg, master):
        cfg = source_cfg(["件名"], [["A"], ["B"], ["C"]])
        counts = etb.process_source(cfg, FakeBacklog(), master, dry_run=False)
        assert counts["created"] == 3


class TestSummaryDisplay:
    """print_summary の未到達だった分岐。"""

    def test_ドライランでスキップとエラーが出る(self, capsys):
        total = etb.new_counts()
        total.update(created=1, skipped=2, error=1)

        etb.print_summary(total, dry_run=True)

        out = capsys.readouterr().out
        assert "スキップ: 2 件" in out
        assert "エラー: 1 件" in out

    def test_実行ログのパスと再開コマンドが出る(self, tmp_path, capsys):
        total = etb.new_counts()
        total.update(created=1, error=1)

        etb.print_summary(total, dry_run=False, log_path=tmp_path / "run_20260101.csv")

        out = capsys.readouterr().out
        assert "実行ログ:" in out
        assert "--resume run_20260101.csv" in out

    def test_エラーが無ければ再開コマンドは出さない(self, tmp_path, capsys):
        total = etb.new_counts()
        total["created"] = 1

        etb.print_summary(total, dry_run=False, log_path=tmp_path / "run.csv")

        out = capsys.readouterr().out
        assert "実行ログ:" in out
        assert "--resume" not in out


class TestConfirmRunDisplay:
    def test_件数が表示される(self, capsys):
        planned = etb.new_counts()
        planned.update(created=5, updated=2, unchanged=1, skipped=3)

        etb.confirm_run([{"name": "S"}], None, assume_yes=True, planned=planned)

        out = capsys.readouterr().out
        assert "作成予定: 5 件 / 更新予定: 2 件" in out
        assert "変更なし: 1 件" in out
        assert "スキップ: 3 件" in out

    def test_算出できなかった場合は件数を出さない(self, capsys):
        etb.confirm_run([{"name": "S"}], None, assume_yes=True, planned=None)

        out = capsys.readouterr().out
        assert "作成予定" not in out
        assert "対象ソース: S" in out

    def test_該当のない項目は出さない(self, capsys):
        planned = etb.new_counts()
        planned["created"] = 1

        etb.confirm_run([{"name": "S"}], None, assume_yes=True, planned=planned)

        out = capsys.readouterr().out
        assert "変更なし" not in out
        assert "再開スキップ" not in out


class TestListMasterFallbacks:
    def test_取得できなかった項目はその旨を出す(self, capsys):
        from mapper import BacklogMaster
        etb.print_master_data(BacklogMaster(project_id=1))
        assert "取得できませんでした" in capsys.readouterr().out
