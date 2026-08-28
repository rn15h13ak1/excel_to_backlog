"""
実行ログと再開のテスト
======================
途中で落ちても何が作られたか分かること、再実行で重複を作らないことを固定する。
"""

import csv

import pytest
from openpyxl import Workbook

import excel_to_backlog as etb
from backlog_client import BacklogAPIError
from mapper import BacklogMaster
from run_log import COMPLETED_ACTIONS, RunLog, load_completed


@pytest.fixture
def excel(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "項番", "件名"
    for i, name in enumerate(["課題A", "課題B", "課題C"], start=2):
        ws[f"A{i}"], ws[f"B{i}"] = i - 1, name
    path = tmp_path / "src.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def source_cfg(excel):
    return {
        "name": "テスト",
        "excel": {"path": excel},
        "issue_mapping": {"issue_type": "タスク", "priority": "中", "summary_col": "件名"},
    }


@pytest.fixture
def master():
    return BacklogMaster(project_id=1, issue_type_map={"タスク": 1}, priority_map={"中": 3})


class FakeClient:
    """N 件目の作成で失敗するクライアント。"""

    def __init__(self, fail_at=None):
        self.fail_at = fail_at
        self.calls = 0

    def create_issue(self, params):
        self.calls += 1
        if self.calls == self.fail_at:
            raise BacklogAPIError("作成に失敗しました", status=400)
        return {"issueKey": f"PROJ-{self.calls}", "summary": params["summary"]}


def read_log(path):
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


class TestRunLog:
    def test_ヘッダーが書かれる(self, tmp_path):
        path = tmp_path / "run.csv"
        with RunLog(path):
            pass
        assert read_log(path) == []
        assert path.read_text(encoding="utf-8-sig").startswith("time,source,row,action")

    def test_1件ごとに追記され即座に読める(self, tmp_path):
        """途中で強制終了しても直前までの内容が残ること。"""
        path = tmp_path / "run.csv"
        with RunLog(path) as log:
            log.record(source="S", row=1, action="created",
                       issue_key="PROJ-1", summary="課題A")
            # まだ with を抜けていないが、flush 済みなので読める
            rows = read_log(path)
            assert len(rows) == 1
            assert rows[0]["issue_key"] == "PROJ-1"

    def test_detail_の改行は1行に畳まれる(self, tmp_path):
        path = tmp_path / "run.csv"
        with RunLog(path) as log:
            log.record(source="S", row=1, action="error", detail="1行目\n  2行目")
        assert read_log(path)[0]["detail"] == "1行目 2行目"


class TestProcessSourceLogging:
    def test_作成した課題が記録される(self, source_cfg, master, tmp_path):
        path = tmp_path / "run.csv"
        with RunLog(path) as log:
            etb.process_source(source_cfg, FakeClient(), master, dry_run=False, run_log=log)

        rows = read_log(path)
        assert [r["action"] for r in rows] == ["created"] * 3
        assert [r["issue_key"] for r in rows] == ["PROJ-1", "PROJ-2", "PROJ-3"]
        assert [r["summary"] for r in rows] == ["課題A", "課題B", "課題C"]

    def test_失敗した行も記録され成功分は残る(self, source_cfg, master, tmp_path):
        path = tmp_path / "run.csv"
        with RunLog(path) as log:
            counts = etb.process_source(
                source_cfg, FakeClient(fail_at=2), master, dry_run=False, run_log=log
            )

        assert counts["created"] == 2
        assert counts["error"] == 1
        actions = [r["action"] for r in read_log(path)]
        assert actions == ["created", "error", "created"]

    def test_ログ未指定でも動作する(self, source_cfg, master):
        counts = etb.process_source(source_cfg, FakeClient(), master, dry_run=False)
        assert counts["created"] == 3


class TestResume:
    def test_処理済みの行は飛ばされる(self, source_cfg, master, tmp_path):
        first = tmp_path / "run1.csv"
        with RunLog(first) as log:
            etb.process_source(
                source_cfg, FakeClient(fail_at=2), master, dry_run=False, run_log=log
            )

        completed = load_completed(first)
        assert completed == {("テスト", "課題A"), ("テスト", "課題C")}

        client = FakeClient()
        counts = etb.process_source(
            source_cfg, client, master, dry_run=False, completed=completed
        )

        # 失敗した課題B だけが再実行される
        assert counts["resumed"] == 2
        assert counts["created"] == 1
        assert client.calls == 1

    def test_成功した実行を再開すると何も作らない(self, source_cfg, master, tmp_path):
        path = tmp_path / "run.csv"
        with RunLog(path) as log:
            etb.process_source(source_cfg, FakeClient(), master, dry_run=False, run_log=log)

        client = FakeClient()
        counts = etb.process_source(
            source_cfg, client, master, dry_run=False, completed=load_completed(path)
        )
        assert counts["resumed"] == 3
        assert client.calls == 0

    def test_エラーとスキップは処理済みに含めない(self):
        assert "error" not in COMPLETED_ACTIONS
        assert "skipped" not in COMPLETED_ACTIONS

    def test_作成後にステータス変更が失敗した行は処理済み扱い(self):
        """課題は作成済みのため、再実行すると重複してしまう。"""
        assert "created_status_failed" in COMPLETED_ACTIONS

    def test_存在しないログを指定するとエラー(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_completed(tmp_path / "ない.csv")
