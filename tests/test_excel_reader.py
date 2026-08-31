"""
ExcelReader のテスト
====================
取り消し線（cell_to_markdown）は継承ロジックの追加・削除・再追加を繰り返して
きた箇所のため、往復した4パターンをすべて固定する。
"""

import pytest
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from excel_reader import ExcelReader, cell_to_markdown, cell_to_str


# ------------------------------------------------------------------
# ヘッダーの一意化
# ------------------------------------------------------------------

class TestDuplicateHeaders:
    """
    ヘッダー名は Excel の表記をそのまま使う。連番などは付けない。
    Excel と見比べたときに名前が食い違うと混乱するため。

    同名の列は左端だけを使い、右側は読み込まない。黙って捨てると
    気づけないので、どの列が使われないかを警告する。
    """

    def test_ヘッダー名をそのまま返す(self, make_excel):
        path = make_excel(["備考", "件名", "備考"], [["B", "課題A", "D"]])
        headers, _ = ExcelReader({"path": str(path)}).read()
        assert headers == ["備考", "件名", "備考"]      # 連番を付けない

    def test_同名の列は左端が使われる(self, make_excel):
        path = make_excel(["備考", "件名", "備考"], [["左の値", "課題A", "右の値"]])
        _, rows = ExcelReader({"path": str(path)}).read()
        assert rows[0]["備考"] == "左の値"

    def test_どの列が無視されるか警告する(self, make_excel, capsys):
        path = make_excel(["項番", "備考", "件名", "備考"], [[1, "B", "課題A", "D"]])
        ExcelReader({"path": str(path)}).read()

        err = capsys.readouterr().err
        assert "同名のヘッダー" in err
        assert "B列を使用" in err
        assert "D列は無視" in err

    def test_3つ以上でもすべて警告する(self, make_excel, capsys):
        path = make_excel(["備考"] * 3, [["A", "B", "C"]])
        ExcelReader({"path": str(path)}).read()
        assert capsys.readouterr().err.count("列は無視") == 2

    def test_重複がなければ警告しない(self, make_excel, capsys):
        path = make_excel(["項番", "件名"], [[1, "課題A"]])
        ExcelReader({"path": str(path)}).read()
        assert "同名のヘッダー" not in capsys.readouterr().err

    def test_col_start_を考慮した列記号を出す(self, make_excel, capsys):
        path = make_excel(["項番", "備考", "備考"], [[1, "B", "C"]])
        ExcelReader({"path": str(path), "col_start": "B"}).read()

        err = capsys.readouterr().err
        assert "B列を使用" in err and "C列は無視" in err

    def test_前後の空白は除去されるため同名になる(self, make_excel, capsys):
        """
        ヘッダーセルの値は cell_to_str が strip する。
        Excel 上で「備考」と「備考 」に見えても、読み込み後は同名になる。
        """
        path = make_excel(["備考", "備考 "], [["左", "右"]])
        headers, rows = ExcelReader({"path": str(path)}).read()

        assert headers == ["備考", "備考"]
        assert rows[0]["備考"] == "左"
        assert "同名のヘッダー" in capsys.readouterr().err

    def test_書式付き行でも左端が優先される(self, make_excel):
        path = make_excel(["備考", "備考"], [["左", "右"]])
        _, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert plain[0]["備考"] == "左"
        assert formatted[0]["備考"] == "左"


# ------------------------------------------------------------------
# セル値の文字列化
# ------------------------------------------------------------------

class TestCellToStr:
    def test_None_は空文字列(self):
        assert cell_to_str(None) == ""

    def test_整数はそのまま(self):
        """openpyxl は整数セルを int で返すため "1.0" にはならない。"""
        assert cell_to_str(1) == "1"

    def test_小数は小数のまま(self):
        assert cell_to_str(0.5) == "0.5"

    def test_前後の空白は除去される(self):
        assert cell_to_str("  値  ") == "値"


# ------------------------------------------------------------------
# 取り消し線 → Markdown
# ------------------------------------------------------------------

class _FakeFont:
    def __init__(self, strike):
        self.strike = strike


class _FakeCell:
    """
    cell_to_markdown() が参照するのは .value と .font.strike だけのため、
    その2つだけを持つ最小のセルを使う。

    openpyxl 3.1 は CellRichText を含むブックの保存に失敗するため、
    実ファイルを往復させる方式は使えない。値そのものは openpyxl の
    CellRichText / TextBlock を使って組み立てるので、読み込み時に得られる
    構造と同じものを検証できる。
    """

    def __init__(self, value, *, strike=False):
        self.value = value
        self.font = _FakeFont(strike)


def _cell(value, *, strike=False):
    """セルレベルの取り消し線を持つ通常セルを作る。"""
    return _FakeCell(value, strike=strike)


def _rich_cell(runs, *, cell_strike=False):
    """
    CellRichText セルを作る。

    runs: [(テキスト, 取り消し線 or None), ...]
          None を渡すと <rPr> を持たないプレーン文字列ランになる
          （openpyxl は書式指定のないランを素の str で返す）。
    """
    blocks = [
        text if struck is None else TextBlock(InlineFont(strike=struck), text)
        for text, struck in runs
    ]
    return _FakeCell(CellRichText(*blocks), strike=cell_strike)


class TestCellToMarkdown:
    """
    継承ロジックが追加・削除・再追加された経緯があるため、
    判断が分かれた4パターンをすべて固定する。
    """

    def test_空セルは空文字列(self):
        assert cell_to_markdown(_cell(None)) == ""

    def test_書式なしのセルはそのまま(self):
        assert cell_to_markdown(_cell("通常テキスト")) == "通常テキスト"

    def test_セル全体の取り消し線は全体を囲む(self):
        """CellRichText でないセル（プレーンテキスト・日付）が対象。"""
        assert cell_to_markdown(_cell("削除済み", strike=True)) == "~~削除済み~~"

    def test_リッチテキストの一部だけ取り消し線(self):
        cell = _rich_cell([("残る", False), ("消える", True)])
        assert cell_to_markdown(cell) == "残る ~~消える~~"

    def test_リッチテキストではランの情報がセルスタイルより優先される(self):
        """
        セルの一部に取り消し線を付けると Excel はセルレベルにも strike=True を
        記録することがある。ランに明示的な strike=True があるときは、
        プレーン文字列ランへ継承してはいけない。
        """
        cell = _rich_cell([("残る", None), ("消える", True)], cell_strike=True)
        assert cell_to_markdown(cell) == "残る ~~消える~~"

    def test_ランに取り消し線がなければセルスタイルを継承する(self):
        """
        「セル全体に取り消し線 → 一部を解除」の操作では、解除したランだけが
        strike=False の TextBlock になり、残りは <rPr> を省略してセルレベルの
        スタイルを継承する。この場合はプレーン文字列ランも取り消し線扱いにする。
        """
        cell = _rich_cell([("消える", None), ("残る", False)], cell_strike=True)
        assert cell_to_markdown(cell) == "~~消える~~ 残る"

    def test_複数行セルは行ごとに取り消し線を適用する(self):
        """Markdown の ~~ は改行をまたげないため行単位で囲む。"""
        cell = _cell("1行目\n2行目", strike=True)
        assert cell_to_markdown(cell) == "~~1行目~~\n~~2行目~~"


# ------------------------------------------------------------------
# 行のフィルタリング
# ------------------------------------------------------------------

class TestFilterRows:
    ROWS = [
        {"ステータス": "対応要", "種別": "バグ"},
        {"ステータス": "完了", "種別": "バグ"},
        {"ステータス": "完了", "種別": "タスク"},
    ]

    def test_条件なしなら全行(self):
        assert ExcelReader.filter_rows(self.ROWS, []) == self.ROWS

    def test_完全一致(self):
        got = ExcelReader.filter_rows(self.ROWS, [{"col_name": "ステータス", "value": "完了"}])
        assert len(got) == 2

    def test_values_は_OR_条件(self):
        got = ExcelReader.filter_rows(
            self.ROWS, [{"col_name": "種別", "values": ["バグ", "タスク"]}]
        )
        assert len(got) == 3

    def test_複数条件は_AND(self):
        got = ExcelReader.filter_rows(self.ROWS, [
            {"col_name": "ステータス", "value": "完了"},
            {"col_name": "種別", "value": "タスク"},
        ])
        assert len(got) == 1

    def test_前方一致(self):
        got = ExcelReader.filter_rows(
            self.ROWS, [{"col_name": "ステータス", "value": "完", "match": "startswith"}]
        )
        assert len(got) == 2

    def test_部分一致(self):
        got = ExcelReader.filter_rows(
            self.ROWS, [{"col_name": "ステータス", "value": "対応", "match": "contains"}]
        )
        assert len(got) == 1

    def test_存在しない列の条件は無視され全行が通る(self):
        """
        この挙動自体は危険だが、実行前に validate_column_references() が
        列名を検証して停止するため、ここへは到達しない。
        挙動を変える場合は検証側とあわせて見直すこと。
        """
        got = ExcelReader.filter_rows(
            self.ROWS, [{"col_name": "存在しない列", "value": "x"}]
        )
        assert len(got) == 3


# ------------------------------------------------------------------
# 設定のバリデーション
# ------------------------------------------------------------------

class TestExcelConfigValidation:
    def test_header_start_row_は1以上(self):
        with pytest.raises(ValueError, match="header_start_row"):
            ExcelReader({"path": "x.xlsx", "header_start_row": 0})

    def test_header_end_row_は_start_以上(self):
        with pytest.raises(ValueError, match="header_end_row"):
            ExcelReader({"path": "x.xlsx", "header_start_row": 3, "header_end_row": 2})

    def test_data_start_row_は_header_end_row_より大きい(self):
        with pytest.raises(ValueError, match="data_start_row"):
            ExcelReader({"path": "x.xlsx", "header_end_row": 3, "data_start_row": 3})

    def test_存在しないファイルは_FileNotFoundError(self, tmp_path):
        reader = ExcelReader({"path": str(tmp_path / "ない.xlsx")})
        with pytest.raises(FileNotFoundError):
            reader.read()


class TestColumnRange:
    """col_start / col_end による読み込み範囲の指定。"""

    def _sheet(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        for col, header in zip("ABCD", ["項番", "件名", "備考", "内部用"]):
            ws[f"{col}1"] = header
        ws["A2"], ws["B2"], ws["C2"], ws["D2"] = 1, "課題A", "メモ", "見せない"
        path = tmp_path / "cols.xlsx"
        wb.save(path)
        return path

    def test_既定では最終列まで読む(self, tmp_path):
        headers, _ = ExcelReader({"path": str(self._sheet(tmp_path))}).read()
        assert headers == ["項番", "件名", "備考", "内部用"]

    def test_col_end_で右端を絞れる(self, tmp_path):
        headers, rows = ExcelReader(
            {"path": str(self._sheet(tmp_path)), "col_end": "C"}
        ).read()
        assert headers == ["項番", "件名", "備考"]
        assert "内部用" not in rows[0]

    def test_col_start_で左端を絞れる(self, tmp_path):
        headers, _ = ExcelReader(
            {"path": str(self._sheet(tmp_path)), "col_start": "B", "col_end": "C"}
        ).read()
        assert headers == ["件名", "備考"]

    def test_小文字の列ラベルも受け付ける(self, tmp_path):
        headers, _ = ExcelReader(
            {"path": str(self._sheet(tmp_path)), "col_start": "b", "col_end": "c"}
        ).read()
        assert headers == ["件名", "備考"]

    def test_col_start_が_col_end_より後ならエラー(self, tmp_path):
        reader = ExcelReader(
            {"path": str(self._sheet(tmp_path)), "col_start": "C", "col_end": "A"}
        )
        with pytest.raises(ValueError, match="col_start"):
            reader.read()


class TestSheetSelection:
    def _book(self, tmp_path):
        wb = Workbook()
        first = wb.active
        first.title = "最初のシート"
        first["A1"], first["A2"] = "件名", "1枚目"
        second = wb.create_sheet("対象シート")
        second["A1"], second["A2"] = "件名", "2枚目"
        path = tmp_path / "sheets.xlsx"
        wb.save(path)
        return path

    def test_省略時は最初のシート(self, tmp_path):
        _, rows = ExcelReader({"path": str(self._book(tmp_path))}).read()
        assert rows[0]["件名"] == "1枚目"

    def test_シート名を指定できる(self, tmp_path):
        _, rows = ExcelReader(
            {"path": str(self._book(tmp_path)), "sheet": "対象シート"}
        ).read()
        assert rows[0]["件名"] == "2枚目"

    def test_存在しないシート名はエラーで候補を示す(self, tmp_path):
        reader = ExcelReader({"path": str(self._book(tmp_path)), "sheet": "無いシート"})
        with pytest.raises(ValueError) as exc:
            reader.read()
        assert "無いシート" in str(exc.value)
        assert "対象シート" in str(exc.value)      # 利用可能な名前を出す


class TestDateCells:
    """日付型セルは "YYYY/MM/DD" 文字列として読む。"""

    def test_datetime_セルを日付文字列にする(self, tmp_path):
        from datetime import datetime
        wb = Workbook()
        ws = wb.active
        ws["A1"], ws["A2"] = "期限", datetime(2025, 1, 5, 10, 30)
        path = tmp_path / "dt.xlsx"
        wb.save(path)

        _, rows = ExcelReader({"path": str(path)}).read()

        assert rows[0]["期限"] == "2025/01/05"

    def test_date_セルも同じ形式(self, tmp_path):
        from datetime import date
        wb = Workbook()
        ws = wb.active
        ws["A1"], ws["A2"] = "期限", date(2025, 12, 31)
        path = tmp_path / "d.xlsx"
        wb.save(path)

        _, rows = ExcelReader({"path": str(path)}).read()

        assert rows[0]["期限"] == "2025/12/31"

    def test_日付セルの取り消し線も行ごとに囲む(self):
        from datetime import date
        assert cell_to_markdown(_cell(date(2025, 1, 5), strike=True)) == "~~2025/01/05~~"


class TestRichTextUnavailable:
    """
    古い openpyxl では rich_text が使えない。警告を出して平文で続行する。
    静かに機能が失われると、取り消し線が反映されない原因が分からない。
    """

    def test_警告を出して平文行を返す(self, tmp_path, monkeypatch, capsys):
        import excel_reader
        monkeypatch.setattr(excel_reader, "_RICH_TEXT_AVAILABLE", False)

        wb = Workbook()
        ws = wb.active
        ws["A1"], ws["A2"] = "件名", "課題A"
        path = tmp_path / "plain.xlsx"
        wb.save(path)

        headers, plain, formatted = ExcelReader({"path": str(path)}).read_with_format()

        assert plain == formatted            # 書式付き行は平文と同じ
        assert "リッチテキスト機能が利用できません" in capsys.readouterr().err
